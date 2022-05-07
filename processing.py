import os
import openpyxl
from openpyxl import Workbook, load_workbook


def write2excel(img_names, pred_labels, opt):
    for i in range(len(img_names)):
        img_names[i] = img_names[i].split('/')[-1].split('.')[0]
    if len(img_names) != len(pred_labels):
        print('图片名列表与结果列表长度不同，请检查！')
        return
    a = os.path.join(opt['evaluation_path'], 'label.xlsx')
    if os.path.exists(a) == True:
        print('当前文件夹下有同名文件，请删除或等待一秒后重试！')
        return
    else:
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet(title='sheet1', index=0)
        data = []
        list1 = img_names
        list2 = pred_labels
        for i in range(0, len(img_names)):
            data.append([list1[i], list2[i]])
        for i in data:
            sheet.append(i)
        wb.save(a)
        print('生成成功！生成的文件名字为：' + a)
        return a


def write_text(filename, string):
    file_path = filename
    file = open(file_path, 'w')
    file.write(string)
    file.close()
